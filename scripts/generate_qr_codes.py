#!/usr/bin/env python3
"""Generate one QR code PNG per row in an Excel file of account/password pairs.

Usage:
    pip install qrcode[pil] openpyxl
    python generate_qr_codes.py accounts.xlsx -o qr_out/

Each QR encodes a single line of plain text:
    Username: <account> Password: <password>

PNG files are named qr_<account>.png (account sanitized for filesystem).
"""

import argparse
import re
import sys
from pathlib import Path

import qrcode
from openpyxl import load_workbook


def sanitize_filename(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "account"


def read_accounts(xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError(f"{xlsx_path} has no header row")

    headers_lower = [str(h).strip().lower() if h is not None else "" for h in header]
    try:
        acc_idx = headers_lower.index("account")
        pw_idx = headers_lower.index("password")
    except ValueError:
        raise ValueError(
            f"Expected columns 'account' and 'password' in the first sheet; "
            f"got headers: {header}"
        )

    for row in rows:
        if row is None:
            continue
        account = row[acc_idx]
        password = row[pw_idx]
        if account is None or password is None:
            continue
        account = str(account).strip()
        password = str(password).strip()
        if not account or not password:
            continue
        yield account, password


def make_qr(payload: str, out_path: Path) -> None:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="Path to the Excel file")
    parser.add_argument(
        "-o", "--out-dir", type=Path, default=Path("qr_codes"),
        help="Output directory for PNGs (default: ./qr_codes)",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} not found", file=sys.stderr)
        return 1

    args.out_dir.mkdir(parents=True, exist_ok=True)

    count = 0
    seen = set()
    for account, password in read_accounts(args.xlsx):
        base = sanitize_filename(account)
        filename = f"qr_{base}.png"
        if filename in seen:
            suffix = 2
            while f"qr_{base}_{suffix}.png" in seen:
                suffix += 1
            filename = f"qr_{base}_{suffix}.png"
        seen.add(filename)

        payload = f"Username: {account} Password: {password}"
        make_qr(payload, args.out_dir / filename)
        count += 1

    print(f"Wrote {count} QR code(s) to {args.out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
